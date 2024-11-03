import pandas as pd
import os
import warnings
from openpyxl import load_workbook, Workbook
import configparser
from datetime import datetime

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')


class ImputationProcessor:
    class ImputationProcessor:
        def __init__(self):
            # Dictionnaire pour mapper les abréviations de mois
            self.mois_abbr = {
                1: 'janv', 2: 'févr', 3: 'mars', 4: 'avr', 5: 'mai',
                6: 'juin', 7: 'juil', 8: 'août', 9: 'sept', 10: 'oct',
                11: 'nov', 12: 'déc'
            }

        def formater_date_mois(self, mois):
            """
            Formate une date ou un numéro de mois en utilisant les abréviations définies dans mois_abbr.

            :param mois: Peut être un numéro de mois (int), une date sous forme de chaîne ('YYYY-MM-DD'),
                         une date déjà en format de date (datetime), ou None
            :return: Une chaîne de caractères avec le mois et l'année (ex: 'oct 2024')
                     ou 'Mois inconnu' si le format est incorrect.
            """
            print(f"[DEBUG] Valeur de 'mois' reçue : {mois} | Type : {type(mois)}")

            if isinstance(mois, int):  # Cas où mois est un entier représentant le mois
                mois_abbr = self.mois_abbr.get(mois, "Mois inconnu")
                return mois_abbr

            elif isinstance(mois, str):  # Cas où mois est une chaîne de caractères
                try:
                    # On suppose que le format attendu est 'YYYY-MM-DD'
                    date_obj = datetime.strptime(mois, '%Y-%m-%d')
                    mois_abbr = self.mois_abbr[date_obj.month]
                    return f"{mois_abbr} {date_obj.year}"
                except ValueError:
                    # Si le format de la chaîne ne correspond pas à 'YYYY-MM-DD'
                    print(f"[ERREUR] Format de date non reconnu pour 'mois' : {mois}")
                    return "Mois inconnu"

            elif isinstance(mois, datetime):  # Cas où mois est déjà un objet datetime
                mois_abbr = self.mois_abbr.get(mois.month, "Mois inconnu")
                return f"{mois_abbr} {mois.year}"

            elif mois is None:  # Cas où mois est None
                print("[ERREUR] 'mois' est None.")
                return "Mois inconnu"

            else:
                # Type inattendu, afficher une erreur et renvoyer "Mois inconnu"
                print(f"[ERREUR] Type inattendu pour 'mois' : {type(mois)}")
                return "Mois inconnu"

    def creer_fichier_excel(self):
        """Crée un fichier Excel vide avec les en-têtes nécessaires s'il n'existe pas ou est corrompu."""
        print(f"Création d'un nouveau fichier Excel à {self.output_file}.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Imputations"

        # Ajouter les en-têtes de colonnes
        headers = [
            "Mois", "Nom", "Technologie", "Tâche", "Livrable",
            "Description détaillée", "Code XAN du projet", "Equivelent en jours"
        ]
        ws.append(headers)

        # Sauvegarder le fichier
        wb.save(self.output_file)

    def verify_excel_file(self):
        """Vérifie que le fichier de sortie est un fichier Excel valide."""
        try:
            # Tente de charger le fichier avec openpyxl pour vérifier qu'il est valide
            load_workbook(self.output_file)
            print("Le fichier est valide et peut être utilisé.")
            return True
        except Exception as e:
            print(f"Le fichier n'est pas valide : {e}")
            return False

    def inserer_donnees(self, input_files):
        """Récupère les données des fichiers d'imputation et les ajoute au fichier existant."""
        # Vérifier et créer le fichier de sortie si nécessaire
        if not os.path.exists(self.output_file) or not self.verify_excel_file():
            self.creer_fichier_excel()

        # Charger le fichier de sortie
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
        except Exception as e:
            raise IOError(f"Erreur lors du chargement du fichier de sortie : {e}")

        # Lire et ajouter les données des fichiers sélectionnés
        for input_file in input_files:
            try:
                df_colonnes = pd.read_excel(input_file, sheet_name='Imputation', header=None, engine='openpyxl')
            except Exception as e:
                raise IOError(f"Erreur lors du chargement de {input_file} : {e}")

            if len(df_colonnes) < 10:
                raise ValueError(f"Le fichier '{input_file}' ne contient pas suffisamment de données.")

            # Extraction des données
            nom = df_colonnes.iloc[1, 1]
            mois = df_colonnes.iloc[0, 1]
            mois = self.formater_date_mois(mois)

            technologie = df_colonnes.iloc[10:, 0].tolist()
            tache = df_colonnes.iloc[10:, 1].tolist()
            livrable = df_colonnes.iloc[10:, 2].tolist()
            description_detaillee = df_colonnes.iloc[10:, 4].tolist()
            code_xan_projet = df_colonnes.iloc[10:, 5].tolist()
            nombre_de_jours = df_colonnes.iloc[10:, 6].tolist()

            for i in range(len(technologie)):
                # Ajouter chaque ligne dans le fichier de sortie
                ws.append([
                    mois, nom, technologie[i], tache[i], livrable[i],
                    description_detaillee[i], code_xan_projet[i], nombre_de_jours[i]
                ])

        # Sauvegarder les données mises à jour dans le fichier de sortie
        try:
            wb.save(self.output_file)
            print("Données sauvegardées dans le fichier de sortie.")
        except Exception as e:
            raise IOError(f"Erreur lors de l'enregistrement des données dans le fichier de sortie : {e}")
